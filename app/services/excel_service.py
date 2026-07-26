"""
Lectura del Excel de casos (hojas PP / PT) y mapeo a los nombres de campo
{{...}} usados en las plantillas jsreport. Ver plan_proyecto.md para el detalle
del mapeo verificado campo por campo.
"""
import re
from pathlib import Path
from typing import Literal

import openpyxl

Hoja = Literal["PP", "PT"]

# MERGEFIELD (jsreport) -> columna real del Excel
PP_FIELD_MAP: dict[str, str] = {
    "N_ID": "N° ID",
    "F_ELABORACIÓN_DE_CONVENIO2": "F. ELABORACIÓN DE CONVENIO2",
    "NOMBRES_COMPLETOS_DEL_ASEGURADO": "NOMBRES COMPLETOS DEL ASEGURADO",
    "PRODUCTO": "PRODUCTO",
    "PÓLIZA": "PÓLIZA",
    "NOMBRE__COMERCIAL": "NOMBRE \nCOMERCIAL",
    "CULTIVO_ASEGURADO": "CULTIVO ASEGURADO",
    "EVENTO_CLIMATICO_": "EVENTO CLIMATICO ",
    "LUGAR_DE_OCURRENCIA_": "LUGAR DE OCURRENCIA ",
    "FECHA_DE_SINIESTRO": "FECHA DE SINIESTRO",
    "NÚMERO_DE_CERTIFICADO": "NÚMERO DE CERTIFICADO",
    "INICIO_DE_VIGENCIA": "INICIO DE VIGENCIA",
    "FIN_DE_VIGENCIA": "FIN DE VIGENCIA",
    "RDTO_ASEGURADO2": "RDTO ASEGURADO2",
    "RENDT_OBTENIDO2": "RENDT OBTENIDO2",
    "CANTIDAD_DIFERENCIA2": "CANTIDAD DIFERENCIA2",
    "DIFERENCIAL2": "DIFERENCIAL2",
    "DIFERENCIAL3": "DIFERENCIAL3",
    "CAPÍTULO_CG": "CAPÍTULO CG",
    "SUMA_ASEGURADA_POR_HA2": "SUMA ASEGURADA POR HA2",
    "NOMBRE_AREA_PAGADA": "NOMBRE AREA PAGADA",
    "NOMBRE_AREA_PAGADA2": "NOMBRE AREA PAGADA2",
    "AREA_PAGADA": "AREA PAGADA",
    "NOMBRE_AREA_NO_PAGADA": "NOMBRE AREA NO PAGADA",
    "AREA_NO_PAGADA": "AREA NO PAGADA",
    "LOTES_POR_PAGAR2": "LOTES POR PAGAR2",
    "INDEMNIZACIÓN_PARCIAL2": "INDEMNIZACIÓN (PARCIAL)2",
    "INDEMNIZACIÓN_PARCIAL__LETRAS": "INDEMNIZACIÓN (PARCIAL) - LETRAS",
    "OBSERVACIONES_PARA_CONVENIO": "OBSERVACIONES PARA CONVENIO",
    "ENDOSATARIO": "ENDOSATARIO",
    "RUC": "RUC",
    "DNI_": "DNI ",
}

PT_FIELD_MAP: dict[str, str] = {
    "N_ID": "N° ID",
    "F_ELABORACIÓN_DE_CONVENIO2": "F. ELABORACIÓN DE CONVENIO2",
    "NOMBRES_COMPLETOS_DEL_ASEGURADO": "NOMBRES COMPLETOS DEL ASEGURADO",
    "PRODUCTO": "PRODUCTO",
    "PÓLIZA": "PÓLIZA",
    "NOMBRE__COMERCIAL": "NOMBRE \nCOMERCIAL",
    "CULTIVO_ASEGURADO": "CULTIVO ASEGURADO",
    "EVENTO_CLIMATICO_": "EVENTO CLIMATICO ",
    "LUGAR_DE_OCURRENCIA_": "LUGAR DE OCURRENCIA ",
    "FECHA_DE_SINIESTRO": "FECHA DE SINIESTRO",
    "FECHA_DE_SINIESTRO2": "FECHA DE SINIESTRO2",
    "NÚMERO_DE_CERTIFICADO": "NÚMERO DE CERTIFICADO",
    "INICIO_DE_VIGENCIA": "INICIO DE VIGENCIA",
    "FIN_DE_VIGENCIA": "FIN DE VIGENCIA",
    "FECHA_PRINCIPAL": "FECHA PRINCIPAL",
    "FECHA_PRINCIPAL2": "FECHA PRINCIPAL2",
    "FECHA_PRINCIPAL3": "FECHA PRINCIPAL3",
    "SUMA_ASEGURADA_POR_HA2": "SUMA ASEGURADA POR HA2",
    "NOMBRE_AREA_NO_PAGADA": "NOMBRE AREA NO PAGADA",
    "AREA_NO_PAGADA": "AREA NO PAGADA",
    "LOTES_POR_PAGAR2": "LOTES POR PAGAR2",
    "NOMBRE_AREA_PAGADA": "NOMBRE AREA PAGADA",
    "NOMBRE_AREA_PAGADA2": "NOMBRE AREA PAGADA2",
    "AREA_PAGADA": "AREA PAGADA",
    "INV_CAPITAL_MESES2": "INV. CAPITAL (MESES)2",
    "DIFERENCIAL2": "DIFERENCIAL2",
    "DIFERENCIAL3": "DIFERENCIAL3",
    "CAPÍTULO_CG": "CAPÍTULO CG",
    "INDEMNIZACIÓN_TOTAL2": "INDEMNIZACIÓN (TOTAL)2",
    "INDEMNIZACIÓN_TOTAL__LETRAS": "INDEMNIZACIÓN (TOTAL) - LETRAS",
    "OBSERVACIONES_PARA_CONVENIO": "OBSERVACIONES PARA CONVENIO",
    "ENDOSATARIO": "ENDOSATARIO",
    "RUC": "RUC",
    "DNI_": "DNI ",
}

FIELD_MAPS: dict[Hoja, dict[str, str]] = {"PP": PP_FIELD_MAP, "PT": PT_FIELD_MAP}


def _clean(valor) -> str:
    if valor is None:
        return ""
    texto = str(valor).strip()
    if re.match(r"^\d+\.0$", texto):
        texto = texto[:-2]
    return texto


class FilaInvalida(Exception):
    pass


def parse_excel(ruta: Path, hoja: Hoja) -> list[dict]:
    """
    Lee la hoja indicada (PP o PT) y devuelve una lista de casos, cada uno
    como dict {MERGEFIELD: valor} + metadata (_nombre_archivo, _fila_excel).
    No incluye ACTA1..6 (eso lo resuelve el servicio de fotos aparte).
    """
    if hoja not in FIELD_MAPS:
        raise ValueError(f"Hoja no soportada: {hoja}")

    field_map = FIELD_MAPS[hoja]

    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    if hoja not in wb.sheetnames:
        raise ValueError(f"El Excel no tiene una hoja llamada '{hoja}'")

    ws = wb[hoja]
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return []

    encabezado = filas[0]
    columnas_faltantes = [
        col for col in field_map.values() if col not in encabezado
    ]
    if columnas_faltantes:
        raise FilaInvalida(
            f"El Excel no tiene estas columnas esperadas para {hoja}: {columnas_faltantes}"
        )

    casos = []
    for i, fila in enumerate(filas[1:], start=2):
        datos_fila = dict(zip(encabezado, fila))
        id_reg = _clean(datos_fila.get("N° ID", ""))
        if not id_reg:
            continue  # fila vacía / de plantilla

        caso = {
            campo: _clean(datos_fila.get(columna, ""))
            for campo, columna in field_map.items()
        }

        caso["_id"] = id_reg
        caso["_nombre_archivo"] = _clean(datos_fila.get("NOMBRE DE ARCHIVO", f"convenio_{id_reg}"))
        caso["_fila_excel"] = i
        casos.append(caso)

    return casos


def parse_excel_todas_las_hojas(ruta: Path) -> dict[Hoja, list[dict]]:
    """
    Lee el Excel completo y devuelve los casos de cada hoja presente
    (PP y/o PT). Si el archivo solo trae una de las dos, no es error:
    simplemente esa hoja queda con lista vacía.
    """
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    resultado: dict[Hoja, list[dict]] = {}
    for hoja in ("PP", "PT"):
        if hoja in wb.sheetnames:
            resultado[hoja] = parse_excel(ruta, hoja)
        else:
            resultado[hoja] = []

    if not resultado["PP"] and not resultado["PT"]:
        raise FilaInvalida(
            "El Excel no tiene hojas 'PP' ni 'PT' con casos, o le faltan las "
            "columnas esperadas. Revisa el archivo."
        )

    return resultado
